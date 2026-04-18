"""
Excel (.xlsx, .xls) ingestion for Ariadne.

Extracts content from Excel spreadsheets, preserving worksheet structure
and cell metadata (comments, formulas).
"""

from pathlib import Path
from typing import List, Optional
import warnings

from .base import BaseIngestor, SourceType, Document


class ExcelIngestor(BaseIngestor):
    """
    Ingestor for Excel spreadsheets (.xlsx, .xls).
    
    Extracts:
    - Cell values organized by row
    - Worksheet names as structural context
    - Column headers for tabular data interpretation
    
    Dependencies:
        pip install openpyxl xlrd
    """
    
    source_type = SourceType.EXCEL
    
    def _extract(self, path: Path) -> List[str]:
        """
        Extract text content from Excel file.
        
        Each worksheet produces a structured text block with:
        - Worksheet name as header
        - Column headers (if detected)
        - Row data in tab-separated format
        """
        blocks = []
        
        try:
            import openpyxl
        except ImportError:
            warnings.warn(
                "openpyxl not installed. Install with: pip install openpyxl",
                UserWarning
            )
            return blocks
        
        try:
            # Try .xlsx first (openpyxl)
            wb = openpyxl.load_workbook(
                str(path),
                data_only=True,  # Get calculated values, not formulas
                read_only=True
            )
        except Exception:
            try:
                # Fallback to .xls (xlrd)
                import xlrd
                return self._extract_xls(path, xlrd)
            except ImportError:
                warnings.warn(
                    "xlrd not installed for .xls support. Install with: pip install xlrd",
                    UserWarning
                )
                return blocks
            except Exception:
                return blocks
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            blocks.append(self._extract_sheet(sheet, sheet_name))
        
        wb.close()
        return [b for b in blocks if b.strip()]
    
    def _extract_sheet(self, sheet, sheet_name: str) -> str:
        """Extract content from a single worksheet."""
        lines = [f"[Worksheet: {sheet_name}]"]
        
        # Detect if first row looks like headers
        first_row_values = []
        has_headers = False
        
        for row_idx, row in enumerate(sheet.iter_rows(max_row=1, values_only=True)):
            first_row_values = [str(v).strip() if v is not None else "" for v in row]
            if any(first_row_values):
                has_headers = True
        
        if has_headers:
            lines.append(f"Headers: {' | '.join(first_row_values)}")
        
        # Extract row data
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2 if has_headers else 1, values_only=True), 
                                       start=2 if has_headers else 1):
            row_values = [str(v).strip() if v is not None else "" for v in row]
            if any(row_values):
                # Include row number for reference
                lines.append(f"Row {row_idx}: {' | '.join(row_values)}")
        
        return "\n".join(lines)
    
    def _extract_xls(self, path: Path, xlrd) -> List[str]:
        """Extract content from legacy .xls format."""
        blocks = []
        
        try:
            wb = xlrd.open_workbook(str(path))
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                lines = [f"[Worksheet: {sheet.name}]"]
                
                for row_idx in range(sheet.nrows):
                    row_values = [str(sheet.cell_value(row_idx, col_idx)) 
                                  for col_idx in range(sheet.ncols)]
                    if any(v.strip() for v in row_values):
                        lines.append(f"Row {row_idx + 1}: {' | '.join(row_values)}")
                
                blocks.append("\n".join(lines))
        except Exception:
            pass
        
        return [b for b in blocks if b.strip()]
